"""
Data Cleaning Script — Russia–Ukraine War Analytics Dashboard
=============================================================
Cleans and merges three source files:
  1. events_table_ukraine_war.csv         — ACLED events (main dataset)
  2. ukraine-war-conflict-tracking-dataset.csv — Monthly aggregates
  3. ukraine_war_conflict_tracking_workbook..xlsx — Excel workbook (same data)

Outputs to:  public/data/
  - events_cleaned.csv          (main map/chart data)
  - monthly_summary_cleaned.csv (monthly KPIs & losses)
  - cleaning_report.txt         (log of cleaning steps)
"""

import csv
import os
import re
import sys
from datetime import datetime

# ── output directory ──────────────────────────────────────────────────────────
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
os.makedirs(OUT_DIR, exist_ok=True)

LOG_LINES = []

def log(msg: str):
    print(msg)
    LOG_LINES.append(msg)

# ─────────────────────────────────────────────────────────────────────────────
# 1.  EVENTS TABLE  (events_table_ukraine_war.csv)
# ─────────────────────────────────────────────────────────────────────────────
EVENTS_SRC = r'C:\Users\SATHYA\Downloads\events_table_ukraine_war.csv'

log("=" * 60)
log("STEP 1: Cleaning events_table_ukraine_war.csv")
log("=" * 60)

EXPECTED_COLS = [
    'event_date', 'event_type', 'sub_event_type',
    'actor1', 'assoc_actor_1', 'inter1',
    'actor2', 'assoc_actor_2', 'inter2',
    'civilian_targeting', 'country',
    'admin_lvl1', 'admin_lvl2', 'admin_lvl3',
    'specific_location',
    'latitude', 'longitude', 'geo_precision',
    'source', 'source_scale',
    'notes', 'fatalities', 'tags',
    'population_1km', 'population_5km',
]

NUMERIC_COLS = {'fatalities', 'latitude', 'longitude', 'geo_precision',
                'inter1', 'inter2', 'population_1km', 'population_5km'}

total_in = 0
total_out = 0
skipped_bad_date = 0
skipped_no_latlon = 0
fixed_numeric = 0
fixed_strings = 0

cleaned_events = []

with open(EVENTS_SRC, 'r', encoding='utf-8-sig', errors='replace') as f:
    reader = csv.DictReader(f)
    actual_cols = reader.fieldnames or []
    log(f"  Source columns ({len(actual_cols)}): {actual_cols}")

    # Column mapping: handle slight name differences
    col_map = {}
    for ec in EXPECTED_COLS:
        for ac in actual_cols:
            if ac.strip().lower() == ec.lower():
                col_map[ec] = ac
                break
        if ec not in col_map:
            col_map[ec] = ec  # fallback (may be missing)

    for row in reader:
        total_in += 1

        # ── 1a. Normalise column names ────────────────────────────────────────
        clean = {}
        for ec, ac in col_map.items():
            val = row.get(ac, '').strip()
            clean[ec] = val

        # ── 1b. Parse & validate date ─────────────────────────────────────────
        raw_date = clean.get('event_date', '')
        parsed_date = None
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                parsed_date = datetime.strptime(raw_date, fmt)
                break
            except ValueError:
                continue
        if parsed_date is None:
            skipped_bad_date += 1
            continue
        clean['event_date'] = parsed_date.strftime('%Y-%m-%d')
        clean['year'] = str(parsed_date.year)
        clean['month'] = parsed_date.strftime('%Y-%m')

        # ── 1c. Validate lat/lon ──────────────────────────────────────────────
        try:
            lat = float(clean.get('latitude', ''))
            lon = float(clean.get('longitude', ''))
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                raise ValueError
        except (ValueError, TypeError):
            skipped_no_latlon += 1
            continue
        clean['latitude'] = str(round(lat, 6))
        clean['longitude'] = str(round(lon, 6))

        # ── 1d. Numeric columns ───────────────────────────────────────────────
        for nc in NUMERIC_COLS - {'latitude', 'longitude'}:
            raw = clean.get(nc, '')
            try:
                val = float(raw)
                clean[nc] = str(int(val)) if nc in {'fatalities', 'geo_precision', 'inter1', 'inter2'} else str(round(val, 2))
                fixed_numeric += 1
            except (ValueError, TypeError):
                clean[nc] = '0' if nc == 'fatalities' else ''

        # ── 1e. String cleaning ───────────────────────────────────────────────
        for sc in ('event_type', 'sub_event_type', 'actor1', 'actor2',
                   'admin_lvl1', 'admin_lvl2', 'admin_lvl3', 'specific_location',
                   'civilian_targeting', 'country'):
            val = clean.get(sc, '').strip()
            # Remove non-printable chars
            val = re.sub(r'[^\x20-\x7E\u00C0-\u024F\u0400-\u04FF]', '', val)
            clean[sc] = val
            fixed_strings += 1

        # ── 1f. Civilian targeting normalise ──────────────────────────────────
        ct = clean.get('civilian_targeting', '').lower()
        clean['civilian_targeting'] = 'Yes' if ct in ('civilian targeting', 'yes', '1', 'true') else 'No'

        # ── 1g. Shorten notes (cap at 500 chars for performance) ──────────────
        notes = clean.get('notes', '')
        clean['notes'] = notes[:500] if len(notes) > 500 else notes

        cleaned_events.append(clean)
        total_out += 1

# Output columns (add derived year/month)
OUT_COLS = EXPECTED_COLS + ['year', 'month']

out_path = os.path.join(OUT_DIR, 'events_cleaned.csv')
with open(out_path, 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=OUT_COLS, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(cleaned_events)

log(f"  Input rows      : {total_in:,}")
log(f"  Skipped (date)  : {skipped_bad_date:,}")
log(f"  Skipped (latlon): {skipped_no_latlon:,}")
log(f"  Output rows     : {total_out:,}")
log(f"  Output file     : {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# 2.  MONTHLY SUMMARY  (ukraine-war-conflict-tracking-dataset.csv)
# ─────────────────────────────────────────────────────────────────────────────
MONTHLY_SRC = r'C:\Users\SATHYA\Downloads\ukraine-war-conflict-tracking-dataset.csv'

log("")
log("=" * 60)
log("STEP 2: Cleaning ukraine-war-conflict-tracking-dataset.csv")
log("=" * 60)

monthly_rows = []
monthly_total_in = 0
monthly_total_out = 0
monthly_skipped = 0

with open(MONTHLY_SRC, 'r', encoding='utf-8-sig', errors='replace') as f:
    reader = csv.DictReader(f)
    monthly_headers = reader.fieldnames or []
    log(f"  Source columns ({len(monthly_headers)}): {monthly_headers}")

    for row in reader:
        monthly_total_in += 1

        # Validate year_month
        ym = row.get('year_month', '').strip()
        try:
            datetime.strptime(ym, '%Y-%m')
        except ValueError:
            monthly_skipped += 1
            continue

        cleaned_row = {'year_month': ym}

        for col in monthly_headers[1:]:
            val = row.get(col, '').strip()
            # Replace NA with empty string
            if val.upper() in ('NA', 'N/A', 'NULL', 'NONE', ''):
                cleaned_row[col] = ''
            else:
                # Try numeric
                try:
                    fval = float(val)
                    # Round financial aid to 3 decimals, integers for counts
                    if 'aid' in col or 'rate' in col:
                        cleaned_row[col] = str(round(fval, 4))
                    else:
                        cleaned_row[col] = str(int(round(fval)))
                except ValueError:
                    cleaned_row[col] = val

        monthly_rows.append(cleaned_row)
        monthly_total_out += 1

monthly_out = os.path.join(OUT_DIR, 'monthly_summary_cleaned.csv')
with open(monthly_out, 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=monthly_headers, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(monthly_rows)

log(f"  Input rows  : {monthly_total_in:,}")
log(f"  Skipped     : {monthly_skipped:,}")
log(f"  Output rows : {monthly_total_out:,}")
log(f"  Output file : {monthly_out}")


# ─────────────────────────────────────────────────────────────────────────────
# 3.  XLSX WORKBOOK — extract Events table sheet as sanity check
# ─────────────────────────────────────────────────────────────────────────────
XLSX_SRC = r'C:\Users\SATHYA\Downloads\ukraine_war_conflict_tracking_workbook..xlsx'

log("")
log("=" * 60)
log("STEP 3: Inspecting xlsx workbook sheets")
log("=" * 60)

try:
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_SRC, read_only=True, data_only=True)
    log(f"  Sheets found: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_count = 0
        for _ in ws.iter_rows():
            rows_count += 1
        log(f"  Sheet '{sheet_name}': ~{rows_count:,} rows")

    wb.close()
    log("  NOTE: xlsx data is covered by the CSVs above. No separate export needed.")
except Exception as e:
    log(f"  WARNING: Could not read xlsx ({e}). CSVs are sufficient.")


# ─────────────────────────────────────────────────────────────────────────────
# 4.  GENERATE QUICK STATS JSON for the dashboard
# ─────────────────────────────────────────────────────────────────────────────
log("")
log("=" * 60)
log("STEP 4: Generating quick-stats.json")
log("=" * 60)

import json
from collections import Counter

event_types: Counter = Counter()
regions: Counter = Counter()
total_fatalities = 0
civilian_fatalities = 0
dates_seen = []

for ev in cleaned_events:
    event_types[ev.get('event_type', 'Unknown')] += 1
    regions[ev.get('admin_lvl1', 'Unknown')] += 1
    try:
        total_fatalities += int(ev.get('fatalities', 0) or 0)
    except ValueError:
        pass
    ct = ev.get('civilian_targeting', 'No')
    try:
        fat = int(ev.get('fatalities', 0) or 0)
        if ct == 'Yes':
            civilian_fatalities += fat
    except ValueError:
        pass
    dates_seen.append(ev.get('event_date', ''))

dates_seen = sorted(d for d in dates_seen if d)
date_min = dates_seen[0] if dates_seen else ''
date_max = dates_seen[-1] if dates_seen else ''

stats = {
    'total_events': total_out,
    'total_fatalities': total_fatalities,
    'civilian_fatalities': civilian_fatalities,
    'date_min': date_min,
    'date_max': date_max,
    'top_event_types': dict(event_types.most_common(10)),
    'top_regions': dict(regions.most_common(10)),
}

stats_path = os.path.join(OUT_DIR, 'quick_stats.json')
with open(stats_path, 'w', encoding='utf-8') as f:
    json.dump(stats, f, indent=2)

log(f"  Total events     : {total_out:,}")
log(f"  Total fatalities : {total_fatalities:,}")
log(f"  Date range       : {date_min} → {date_max}")
log(f"  Output file      : {stats_path}")


# ─────────────────────────────────────────────────────────────────────────────
# 5.  WRITE CLEANING REPORT
# ─────────────────────────────────────────────────────────────────────────────
report_path = os.path.join(OUT_DIR, 'cleaning_report.txt')
with open(report_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(LOG_LINES))

log("")
log(f"✅  All done. Report saved to: {report_path}")
